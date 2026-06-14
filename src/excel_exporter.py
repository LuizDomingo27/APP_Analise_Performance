import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def exportar_xlsx_executivo(df_base, df_detalhe, df_eficiencia):
    """
    Gera um buffer de bytes contendo as planilhas formatadas no modelo executivo.
    
    Camadas e regras de negócio:
    - Separação da camada de visualização (Streamlit) e camada de dados.
    - Estilo executivo corporativo: Cabeçalhos azul-marinho, linhas de grade visíveis, auto-ajuste de colunas.
    - Formatação numérica nativa de porcentagens (valores divididos por 100 e formatados como '0.0%').
    - Formatação de números com separadores de milhares.
    - Realce visual suave baseado nos status de risco e performance.
    """
    buf = io.BytesIO()
    
    # Criar cópias dos DataFrames para não alterar a exibição no Streamlit
    base = df_base.copy()
    detalhe = df_detalhe.copy() if df_detalhe is not None and not df_detalhe.empty else None
    eficiencia = df_eficiencia.copy() if df_eficiencia is not None and not df_eficiencia.empty else None
    
    # Dropar colunas auxiliares HTML se existirem
    cols_to_drop = ["PERF_COR", "PERF_PILL"]
    if detalhe is not None:
        for c in cols_to_drop:
            if c in detalhe.columns:
                detalhe = detalhe.drop(columns=[c])
                
    # Renomear colunas para termos executivos profissionais
    map_base = {
        "OFICINA": "Oficina",
        "COLABORADORES": "Colaboradores",
        "CAP_DIA": "Capacidade/Dia",
        "CAP_SEMANAL": "Capacidade Semanal",
        "TOTAL_ORDENS": "Total de Ordens",
        "TOTAL_PECAS": "Total de Peças (Saldo)",
        "TOTAL_MINUTOS": "Minutos Pendentes",
        "CAP_NO_PRAZO": "Capacidade no Prazo",
        "DIAS_NECESSARIOS": "Dias Necessários",
        "SEMANAS_NECESSARIAS": "Semanas Necessárias",
        "FOLGA_DEFICIT": "Folga/Déficit (Saldo)",
        "EFICIENCIA_PCT": "Consumo da Capacidade",
        "PROBABILIDADE_PCT": "Probabilidade",
        "STATUS": "Status da Entrega",
        "ordens_na_meta": "Ordens na Meta",
        "pct_ordens_na_meta": "Performance da MP",
        "efic_media": "Eficiência Média"
    }
    
    map_detalhe = {
        "ORDEM_MESTRE": "Ordem Mestre",
        "OFICINA": "Oficina",
        "MP": "Matéria-Prima",
        "QTD_PECAS": "Qtd Solicitada",
        "QTD_ENTREGUE": "Qtd Entregue",
        "QTD_PENDENTE": "Qtd Pendente (Saldo)",
        "PCT_ENTREGUE": "% Entregue",
        "MINUTOS": "Minutos Solicitados",
        "MIN_PENDENTE": "Minutos Pendentes",
        "PRAZO_DIAS": "Prazo (Dias)",
        "ENVIO": "Data Envio",
        "DEAD_LINE": "Prazo Limite",
        "PERF_STATUS": "Performance"
    }
    
    map_eficiencia = {
        "OFICINA": "Oficina",
        "ordens_na_meta": "Ordens na Meta",
        "total_ordens": "Total de Ordens",
        "pct_ordens_na_meta": "% Ordens na Meta",
        "efic_media": "Eficiência Média"
    }
    
    # Identificar colunas percentuais para dividir por 100 e aplicar formatação de % nativa
    for pct_col in ["EFICIENCIA_PCT", "PROBABILIDADE_PCT", "pct_ordens_na_meta", "efic_media"]:
        if pct_col in base.columns:
            base[pct_col] = pd.to_numeric(base[pct_col], errors="coerce").fillna(0) / 100.0
            
    if detalhe is not None and "PCT_ENTREGUE" in detalhe.columns:
        detalhe["PCT_ENTREGUE"] = pd.to_numeric(detalhe["PCT_ENTREGUE"], errors="coerce").fillna(0) / 100.0
        
    if eficiencia is not None:
        for pct_col in ["pct_ordens_na_meta", "efic_media"]:
            if pct_col in eficiencia.columns:
                eficiencia[pct_col] = pd.to_numeric(eficiencia[pct_col], errors="coerce").fillna(0) / 100.0
                
    # Renomear as colunas
    base = base.rename(columns=map_base)
    if detalhe is not None:
        from src.utils import fmtDate
        if "ENVIO" in detalhe.columns:
            detalhe["ENVIO"] = detalhe["ENVIO"].apply(fmtDate)
        if "DEAD_LINE" in detalhe.columns:
            detalhe["DEAD_LINE"] = detalhe["DEAD_LINE"].apply(fmtDate)
        detalhe = detalhe.rename(columns=map_detalhe)
    if eficiencia is not None:
        eficiencia = eficiencia.rename(columns=map_eficiencia)
        
    # Escrever para o Excel usando o pandas e openpyxl
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        base.to_excel(w, sheet_name="BASE CONSOLIDADA", index=False)
        if detalhe is not None:
            detalhe.to_excel(w, sheet_name="DETALHE ORDENS", index=False)
        if eficiencia is not None:
            eficiencia.to_excel(w, sheet_name="EFICIENCIA POR OFICINA", index=False)
            
        # Obter o workbook para aplicar a estilização
        wb = w.book
        
        # Estilos reutilizáveis
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Segoe UI", size=11, bold=False, color="333333")
        
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Navy azul executivo
        fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid") # Zebra striping
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Cores para alertas (soft)
        # Verde suave
        fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        font_ok = Font(name="Segoe UI", size=11, bold=True, color="375623")
        
        # Amarelo/Laranja suave
        fill_warn = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        font_warn = Font(name="Segoe UI", size=11, bold=True, color="7F6000")
        
        # Vermelho suave
        fill_danger = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        font_danger = Font(name="Segoe UI", size=11, bold=True, color="C65911")
        
        # Bordas
        thin_side = Side(style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Formatos de número excel
        fmt_int = "#,##0"
        fmt_dec = "#,##0.0"
        fmt_pct = "0.0%"
        
        for name in wb.sheetnames:
            ws = wb[name]
            
            # Garantir que as linhas de grade estejam visíveis
            ws.views.sheetView[0].showGridLines = True
            
            # Estilizar cabeçalho (Linha 1)
            ws.row_dimensions[1].height = 28
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border_all
                
            # Estilizar dados (Linha 2 em diante)
            for row_idx in range(2, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 20
                
                # Zebra striping
                row_fill = fill_zebra if row_idx % 2 == 0 else fill_white
                
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = font_data
                    cell.fill = row_fill
                    cell.border = border_all
                    
                    # Alinhamento e formatação padrão dependendo do tipo de dado
                    val = cell.value
                    col_header = ws.cell(row=1, column=col_idx).value
                    
                    # Determinar Alinhamento e Formato por Coluna
                    if isinstance(val, (int, float)) and not pd.isna(val):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Formatar porcentagens
                        if "%" in col_header or col_header in ["Consumo da Capacidade", "Probabilidade", "Performance da MP", "Eficiência Média"]:
                            cell.number_format = fmt_pct
                        # Formatar decimais (dias, semanas)
                        elif "Necessários" in col_header or "Semanas" in col_header:
                            cell.number_format = fmt_dec
                        # Inteiros normais
                        else:
                            cell.number_format = fmt_int
                            
                    elif isinstance(val, str) or val is None:
                        # Centralizar datas ou colunas de data/prazo
                        if "Data" in col_header or "Limite" in col_header or "Prazo" in col_header or col_header in ["Status da Entrega", "Performance"]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            
                    # Regras de coloração condicional de Status / Risco
                    # 1. Planilha BASE CONSOLIDADA (Status da Entrega)
                    if col_header == "Status da Entrega" and isinstance(val, str):
                        if "viavel" in val.lower() or "viável" in val.lower():
                            cell.fill = fill_ok
                            cell.font = font_ok
                        elif "risco moderado" in val.lower() or "atenção" in val.lower() or "atencao" in val.lower():
                            cell.fill = fill_warn
                            cell.font = font_warn
                        elif "alto risco" in val.lower() or "abaixo da meta" in val.lower() or "insuficiente" in val.lower():
                            cell.fill = fill_danger
                            cell.font = font_danger
                            
                    # 2. Planilha DETALHE ORDENS (Performance)
                    if col_header == "Performance" and isinstance(val, str):
                        if "aceitavel" in val.lower() or "aceitável" in val.lower() or "ok" in val.lower():
                            cell.fill = fill_ok
                            cell.font = font_ok
                        elif "atencao" in val.lower() or "atenção" in val.lower():
                            cell.fill = fill_warn
                            cell.font = font_warn
                        elif "abaixo da meta" in val.lower() or "crítico" in val.lower() or "critico" in val.lower():
                            cell.fill = fill_danger
                            cell.font = font_danger
                            
            # Ajustar automaticamente a largura das colunas
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    # Se tiver valor, calcula o tamanho
                    if cell.value is not None:
                        val_str = str(cell.value)
                        # Se for float formatado como pct, ex: 0.75 -> "75.0%" que tem 5 chars
                        if cell.number_format == fmt_pct and isinstance(cell.value, (int, float)):
                            val_str = f"{cell.value * 100:.1f}%"
                        max_len = max(max_len, len(val_str))
                
                # Definir largura mínima e adicionar margem
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
    buf.seek(0)
    return buf.getvalue()


def exportar_recebimento_xlsx_executivo(df_rec_of, c_om, c_qtd, c_min, c_dia):
    """
    Gera um buffer de bytes contendo as planilhas formatadas no modelo executivo
    para os dados de recebimento/entregas de uma oficina.
    """
    import io
    buf = io.BytesIO()
    
    # Criar cópia e formatar coluna de data
    from src.utils import fmtDate
    df_copy = df_rec_of.copy()
    if c_dia in df_copy.columns:
        df_copy[c_dia] = df_copy[c_dia].apply(fmtDate)
        
    df_renomeado = df_copy.rename(columns={
        c_om: "Ordem Mestre",
        c_qtd: "Peças Entregues",
        c_min: "Minutos",
        c_dia: "Data de Entrega"
    })
    
    cols = ["Ordem Mestre", "Peças Entregues", "Minutos", "Data de Entrega"]
    df_to_export = df_renomeado[[c for c in cols if c in df_renomeado.columns]].copy()
    
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_to_export.to_excel(w, sheet_name="RECEBIMENTO", index=False)
        
        wb = w.book
        ws = wb["RECEBIMENTO"]
        
        # Estilos executivos
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Segoe UI", size=11, bold=False, color="333333")
        
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_side = Side(style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        fmt_int = "#,##0"
        
        ws.views.sheetView[0].showGridLines = True
        
        # Cabeçalho
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        # Linhas de dados
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            fill = fill_zebra if row_idx % 2 == 0 else fill_white
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                cell.fill = fill
                cell.border = border_all
                
                col_name = df_to_export.columns[col_idx - 1]
                if col_name in ["Peças Entregues", "Minutos"]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    try:
                        cell.value = float(cell.value) if cell.value is not None else 0
                    except:
                        pass
                    cell.number_format = fmt_int
                elif col_name == "Data de Entrega":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        # Ajuste de largura
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    buf.seek(0)
    return buf.getvalue()

